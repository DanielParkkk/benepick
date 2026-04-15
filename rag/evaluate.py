"""
베네픽 RAG 평가 스크립트 (evaluate.py)
─────────────────────────────────────
평가 지표: faithfulness, answer_relevancy
ground_truth 없이 돌아가는 2개로 구성
사용법: python evaluate.py
"""

import gc
import os
import json
import numpy as np
import pandas as pd
from datetime import datetime
from dotenv import load_dotenv

# 파이프라인 임포트
from pipeline import rerank, crag_quality_check, generate_answer, get_searcher
from langchain_ollama import ChatOllama

# Ragas — ragas 0.4.3 legacy evaluate() API 사용
# Faithfulness/AnswerRelevancy는 ragas.metrics._* (구버전) 경로에서 import해야
# ragas.metrics.collections.Faithfulness는 SimpleBaseMetric이라 Metric 체크 실패
from ragas import evaluate
from ragas.metrics._faithfulness import Faithfulness
from ragas.metrics._answer_relevance import AnswerRelevancy
from ragas.llms import LangchainLLMWrapper
from ragas.embeddings import LangchainEmbeddingsWrapper
from ragas.run_config import RunConfig
from langchain_huggingface import HuggingFaceEmbeddings
from datasets import Dataset

load_dotenv()

# ── 평가용 질문 (카테고리 태그 포함) ──────────────────────────
TEST_QUESTIONS = [
    # (질문, 카테고리)
    ("서울 청년 월세 지원 받을 수 있나요?",             "청년"),
    ("취업 준비생이 받을 수 있는 훈련비 지원이 뭐가 있나요?", "청년"),
    ("청년 창업 지원금 신청 방법 알려주세요",            "청년"),
    ("대학생이 받을 수 있는 장학금 종류",               "청년"),
    ("청년 주거 지원 정책 알려줘",                     "청년"),
    ("만 34세 미만 청년 고용 지원 프로그램",            "청년"),
    ("청년 내일채움공제 조건",                         "청년"),

    ("65세 이상 혼자 사는 노인 지원 정책",              "노인"),
    ("노인 돌봄 서비스 신청 방법",                     "노인"),
    ("기초연금 받으려면 어떻게 해야 하나요?",            "노인"),
    ("노인 의료비 지원 제도",                          "노인"),

    ("장애인 취업 지원 프로그램 알려줘",                "장애인"),
    ("장애인 의료비 지원 혜택",                        "장애인"),
    ("장애인 활동 지원 서비스 신청 방법",               "장애인"),

    ("기초생활수급자 혜택 뭐가 있어요?",                "저소득"),
    ("긴급복지지원 신청 조건",                         "저소득"),
    ("차상위계층이 받을 수 있는 지원",                  "저소득"),
    ("생계급여 받으려면 소득 기준이 어떻게 되나요?",     "저소득"),

    ("출산 지원금 얼마나 받을 수 있어요?",              "출산·육아"),
    ("어린이집 보육료 지원 신청",                      "출산·육아"),
    ("한부모 가정 아동 양육비 지원",                    "출산·육아"),
    ("육아휴직 급여 얼마나 받나요?",                   "출산·육아"),

    ("다문화 가정 한국어 교육 지원",                   "다문화"),
    ("결혼이민자 복지 서비스",                         "다문화"),

    ("실직한 30대 신청할 수 있는 복지 정책",            "고용·실업"),
    ("국민취업지원제도 신청 방법",                     "고용·실업"),
    ("소상공인 자금 지원 받을 수 있는 조건",            "고용·실업"),

    ("전세 사기 피해자 지원 정책",                     "주거"),
    ("주거급여 신청 자격",                             "주거"),

    ("암 환자 의료비 지원 방법",                       "의료"),
]

# 이전 실험 기준값 (비교용 — 구 토크나이저 + 구 프롬프트)
BASELINE = {
    "설정":             "BM25 하이브리드 (구 토크나이저 + 구 프롬프트)",
    "Faithfulness":     1.000,
    "Answer Relevancy": 0.441,
    "평균":             0.7205,
    "비고":             "이전 (채택 기준)",
}


# ── RAG 파이프라인 실행 ────────────────────────────────────────
def run_rag_pipeline(question: str, searcher, llm) -> dict | None:
    """기존 파이프라인 그대로 돌리고 Ragas 형식으로 반환"""
    try:
        results   = searcher.search(question, top_k=25, alpha=0.6)
        if not results:
            return None
        reranked   = rerank(question, results, top_k=5)
        final_docs = crag_quality_check(question, reranked)
        answer     = generate_answer(question, final_docs, lang_code="ko")
        # evidence_text가 NaN(float)이면 str 변환, None이면 제외
        contexts = [
            str(d["evidence_text"])
            for d in final_docs
            if d.get("evidence_text") is not None and str(d.get("evidence_text")) != "nan"
        ]
        return {
            "question": question,
            "answer":   answer,
            "contexts": contexts if contexts else ["관련 정책 정보를 찾을 수 없습니다."],
        }
    except Exception as e:
        print(f"[ERROR] '{question[:25]}...' 처리 중 오류: {e}")
        return None


# ── 점수 안전 추출 헬퍼 ───────────────────────────────────────
def safe_score(val) -> float:
    if isinstance(val, list):
        return float(np.nanmean([v for v in val if v is not None]))
    if val is None:
        return 0.0
    try:
        return float(val)
    except Exception:
        return 0.0


# ── Excel 저장 (스타일 적용) ──────────────────────────────────
def save_excel(df_detail: pd.DataFrame, df_category: pd.DataFrame,
               df_summary: pd.DataFrame, path: str):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # ① 요약 비교 시트
        df_summary.to_excel(writer, sheet_name="요약 비교", index=False)
        ws = writer.sheets["요약 비교"]
        _style_summary(ws, df_summary)

        # ② 카테고리별 시트
        df_category.to_excel(writer, sheet_name="카테고리별", index=False)
        ws2 = writer.sheets["카테고리별"]
        _style_category(ws2, df_category)

        # ③ 질문별 상세 시트
        df_detail.to_excel(writer, sheet_name="질문별 상세", index=False)
        ws3 = writer.sheets["질문별 상세"]
        _auto_width(ws3)

    print(f"  Excel 저장 완료: {path}")


def _style_summary(ws, df):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL  = PatternFill("solid", fgColor="2F4F8F")
    BASE_FILL    = PatternFill("solid", fgColor="D9D9D9")
    CURR_FILL    = PatternFill("solid", fgColor="C6EFCE")   # 연두색
    WHITE_FONT   = Font(bold=True, color="FFFFFF")
    BOLD         = Font(bold=True)
    CENTER       = Alignment(horizontal="center", vertical="center")
    THIN         = Side(style="thin")
    BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    col_widths = [45, 16, 18, 10, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22

    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill      = HEADER_FILL
        cell.font      = WHITE_FONT
        cell.alignment = CENTER
        cell.border    = BORDER

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        fill = CURR_FILL if row_idx == ws.max_row else BASE_FILL
        ws.row_dimensions[row_idx].height = 20
        for cell in row:
            cell.fill      = fill
            cell.font      = BOLD if row_idx == ws.max_row else Font()
            cell.alignment = CENTER
            cell.border    = BORDER


def _style_category(ws, df):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="2F4F8F")
    WHITE_FONT  = Font(bold=True, color="FFFFFF")
    CENTER      = Alignment(horizontal="center", vertical="center")
    THIN        = Side(style="thin")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    col_widths = [15, 10, 18, 18, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = WHITE_FONT
        cell.alignment = CENTER
        cell.border    = BORDER

    # 점수에 따라 셀 색상 (0.7↑ 연두, 0.4↓ 연분홍)
    from openpyxl.styles import PatternFill
    GREEN = PatternFill("solid", fgColor="C6EFCE")
    RED   = PatternFill("solid", fgColor="FFC7CE")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = CENTER
            cell.border    = BORDER
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"
                if cell.value >= 0.7:
                    cell.fill = GREEN
                elif cell.value < 0.4:
                    cell.fill = RED


def _auto_width(ws):
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


# ── 메인 평가 함수 ────────────────────────────────────────────
def run_evaluation():
    print("=" * 60)
    print("베네픽 RAG 평가 시작")
    print(f"평가 질문: {len(TEST_QUESTIONS)}개")
    print("=" * 60)

    # ── 1단계: 파이프라인 실행 (검색 + 답변 수집) ──────────────
    print("\n[1/4] 파이프라인 초기화 중...")
    searcher = get_searcher()   # pipeline.py 전역 싱글턴 — CRAG 폴백도 같은 인스턴스 사용
    llm      = ChatOllama(model="qwen3.5:4b", temperature=0.3)

    print(f"\n[2/4] RAG 파이프라인 실행 중...")
    dataset_rows = []
    categories   = []

    for i, (question, category) in enumerate(TEST_QUESTIONS, 1):
        print(f"  ({i:2d}/{len(TEST_QUESTIONS)}) [{category}] {question[:35]}...")
        result = run_rag_pipeline(question, searcher, llm)
        if result:
            dataset_rows.append(result)
            categories.append(category)

    success = len(dataset_rows)
    failed  = len(TEST_QUESTIONS) - success
    print(f"\n  완료: {success}개 성공 / {failed}개 실패")

    if not dataset_rows:
        print("평가 가능한 데이터가 없습니다!")
        return

    # ── 중간 결과 저장 (Ragas 실패해도 파이프라인 결과는 보존) ──
    interim_path = f"interim_results_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
    with open(interim_path, "w", encoding="utf-8") as f:
        json.dump({"rows": dataset_rows, "categories": categories}, f, ensure_ascii=False, indent=2)
    print(f"  중간 결과 저장: {interim_path}")

    # ── 메모리 해제 (BGE-M3 + reranker VRAM 확보 후 Ragas 로딩) ─
    print("\n  메모리 정리 중...")
    gc.collect()
    try:
        import torch
        torch.cuda.empty_cache()
    except Exception:
        pass

    # ── 2단계: Ragas 평가 (가벼운 임베딩 모델 사용) ───────────
    # BGE-M3를 또 로딩하면 OOM → multilingual-MiniLM 사용 (평가 전용)
    print("\n[3/4] Ragas 평가 실행 중...")
    try:
        import warnings
        warnings.filterwarnings("ignore", category=DeprecationWarning)
        ragas_llm = LangchainLLMWrapper(llm)
        print("  LLM 래퍼 초기화 완료")
        ragas_emb = LangchainEmbeddingsWrapper(
            HuggingFaceEmbeddings(
                model_name="sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2",
                model_kwargs={"device": "cpu"},
            )
        )
        print("  임베딩 모델 로드 완료")
        dataset = Dataset.from_list(dataset_rows)
        print(f"  데이터셋 구성 완료: {len(dataset_rows)}개 행")

        # Faithfulness() / AnswerRelevancy() — 인자 없이 생성, llm/embeddings는 evaluate()가 주입
        print("  Ragas evaluate() 시작 (시간이 오래 걸릴 수 있습니다)...")
        run_cfg = RunConfig(
            timeout=180,      # Ollama 응답 대기 최대 3분
            max_retries=2,
            max_workers=1,    # 직렬 실행 — 병렬 OOM/타임아웃 방지
        )
        result = evaluate(
            dataset=dataset,
            metrics=[Faithfulness(), AnswerRelevancy()],
            llm=ragas_llm,
            embeddings=ragas_emb,
            raise_exceptions=False,
            run_config=run_cfg,
        )
        print("  Ragas evaluate() 완료")
    except Exception as e:
        import traceback
        print(f"\n[ERROR] Ragas 평가 중 오류 발생: {e}")
        traceback.print_exc()
        print(f"\n파이프라인 결과는 {interim_path} 에 저장되어 있습니다.")
        return

    df_ragas = result.to_pandas()
    print(f"  결과 컬럼: {list(df_ragas.columns)}")

    # 점수 추출
    faith_scores = [safe_score(v) for v in df_ragas["faithfulness"]]
    relev_scores = [safe_score(v) for v in df_ragas["answer_relevancy"]]

    faith_avg = round(float(np.nanmean(faith_scores)), 4)
    relev_avg = round(float(np.nanmean(relev_scores)), 4)
    total_avg = round((faith_avg + relev_avg) / 2, 4)

    print("\n" + "=" * 60)
    print("평가 결과")
    print("=" * 60)
    print(f"  Faithfulness      (환각 방지):   {faith_avg:.4f}")
    print(f"  Answer Relevancy  (답변 관련성): {relev_avg:.4f}")
    print(f"  평균:                            {total_avg:.4f}")

    # ── 결과 DataFrame 구성 ──────────────────────────────────

    # ① 질문별 상세
    df_detail = pd.DataFrame({
        "카테고리":          categories,
        "질문":             [r["question"] for r in dataset_rows],
        "Faithfulness":     [round(s, 4) for s in faith_scores],
        "Answer Relevancy": [round(s, 4) for s in relev_scores],
        "평균":             [round((f + a) / 2, 4) for f, a in zip(faith_scores, relev_scores)],
        "답변":             [r["answer"] for r in dataset_rows],
    })

    # ② 카테고리별 집계
    df_detail_tmp = df_detail.copy()
    df_category = (
        df_detail_tmp
        .groupby("카테고리")
        .agg(
            질문수=("질문", "count"),
            Faithfulness=("Faithfulness", "mean"),
            **{"Answer Relevancy": ("Answer Relevancy", "mean")},
            평균=("평균", "mean"),
        )
        .reset_index()
        .round(4)
        .sort_values("평균", ascending=False)
    )

    # ③ 요약 비교 (이전 vs 현재)
    df_summary = pd.DataFrame([
        {
            "설정":             BASELINE["설정"],
            "Faithfulness":     BASELINE["Faithfulness"],
            "Answer Relevancy": BASELINE["Answer Relevancy"],
            "평균":             BASELINE["평균"],
            "비고":             BASELINE["비고"],
        },
        {
            "설정":             "BM25 하이브리드 (Kiwi 형태소 + 개선 프롬프트)",
            "Faithfulness":     faith_avg,
            "Answer Relevancy": relev_avg,
            "평균":             total_avg,
            "비고":             "현재",
        },
    ])

    # ── Excel 저장 ───────────────────────────────────────────
    print("\n[4/4] Excel 저장 중...")
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M")
    excel_path = f"실험결과_Ragas_{timestamp}.xlsx"
    save_excel(df_detail, df_category, df_summary, excel_path)

    # 요약 JSON
    summary = {
        "timestamp":        timestamp,
        "total_questions":  success,
        "faithfulness":     faith_avg,
        "answer_relevancy": relev_avg,
        "average":          total_avg,
        "vs_baseline": {
            "faithfulness_delta":     round(faith_avg - BASELINE["Faithfulness"], 4),
            "answer_relevancy_delta": round(relev_avg - BASELINE["Answer Relevancy"], 4),
            "average_delta":          round(total_avg - BASELINE["평균"], 4),
        },
    }
    json_path = f"실험결과_Ragas_{timestamp}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(f"  JSON 저장 완료: {json_path}")

    # 최종 출력
    print("\n" + "=" * 60)
    print("이전 대비 변화")
    print("=" * 60)
    d_faith = summary["vs_baseline"]["faithfulness_delta"]
    d_relev = summary["vs_baseline"]["answer_relevancy_delta"]
    d_avg   = summary["vs_baseline"]["average_delta"]
    print(f"  Faithfulness:     {BASELINE['Faithfulness']:.4f} → {faith_avg:.4f}  ({'+' if d_faith>=0 else ''}{d_faith:.4f})")
    print(f"  Answer Relevancy: {BASELINE['Answer Relevancy']:.4f} → {relev_avg:.4f}  ({'+' if d_relev>=0 else ''}{d_relev:.4f})")
    print(f"  평균:             {BASELINE['평균']:.4f} → {total_avg:.4f}  ({'+' if d_avg>=0 else ''}{d_avg:.4f})")
    print(f"\n결과 파일: {excel_path}")

    return summary


if __name__ == "__main__":
    run_evaluation()

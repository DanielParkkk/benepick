# -*- coding: utf-8 -*-
"""
실험평가 성능표 Excel 생성 스크립트
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import json
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── 색상 팔레트 ──────────────────────────────────────────────
CLR_HEADER_DARK   = "1F4E79"  # 진한 파랑 (최상단 헤더)
CLR_HEADER_MID    = "2E75B6"  # 중간 파랑 (섹션 헤더)
CLR_HEADER_LIGHT  = "BDD7EE"  # 연한 파랑 (서브 헤더)
CLR_GOOD          = "C6EFCE"  # 연두 (품질 양호)
CLR_WARN          = "FFEB9C"  # 노랑 (조건 완화)
CLR_BAD           = "FFC7CE"  # 연빨강 (폴백)
CLR_SECTION       = "DEEAF1"  # 섹션 구분 배경
CLR_WHITE         = "FFFFFF"
CLR_GRAY          = "F2F2F2"

FONT_WHITE  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
FONT_DARK   = Font(name="맑은 고딕", bold=True, color="1F4E79", size=10)
FONT_BODY   = Font(name="맑은 고딕", size=9)
FONT_BODY_B = Font(name="맑은 고딕", bold=True, size=9)

ALIGN_C  = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_R  = Alignment(horizontal="right",  vertical="center")

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bottom():
    thin = Side(style="thin",  color="BFBFBF")
    thick = Side(style="medium", color="2E75B6")
    return Border(left=thin, right=thin, top=thin, bottom=thick)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def style_header(ws, row, col, value, bg=CLR_HEADER_DARK, font=None, align=ALIGN_C, colspan=1, rowspan=1):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(bg)
    cell.font = font or FONT_WHITE
    cell.alignment = align
    cell.border = thin_border()
    if colspan > 1 or rowspan > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row+rowspan-1, end_column=col+colspan-1
        )
    return cell


def style_cell(ws, row, col, value, bg=CLR_WHITE, font=None, align=ALIGN_C, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(bg)
    cell.font = font or FONT_BODY
    cell.alignment = align
    cell.border = thin_border()
    if number_format:
        cell.number_format = number_format
    return cell


# ════════════════════════════════════════════════════════════
# 시트 1: 파이프라인 인풋/아웃풋 결과
# ════════════════════════════════════════════════════════════
def build_sheet1(wb, rows):
    ws = wb.active
    ws.title = "파이프라인 실행 결과"
    ws.sheet_view.showGridLines = False

    # ── 타이틀 ──
    ws.merge_cells("A1:K1")
    title = ws["A1"]
    title.value = "베네픽(BenePick) RAG 파이프라인 — 질문별 실행 결과 (2026-04-07)"
    title.font = Font(name="맑은 고딕", bold=True, size=13, color="FFFFFF")
    title.fill = fill(CLR_HEADER_DARK)
    title.alignment = ALIGN_C
    title.border = thin_border()
    ws.row_dimensions[1].height = 28

    # ── 서브타이틀 ──
    ws.merge_cells("A2:K2")
    sub = ws["A2"]
    sub.value = "모델: gemma3:1b (Ollama)  |  검색: BM25(40%) + BGE-M3 Dense(60%)  |  Reranker: bge-reranker-v2-m3  |  alpha=0.6  |  질문 수: 15개"
    sub.font = Font(name="맑은 고딕", size=9, color="1F4E79")
    sub.fill = fill(CLR_SECTION)
    sub.alignment = ALIGN_C
    sub.border = thin_border()
    ws.row_dimensions[2].height = 18

    # ── 컬럼 헤더 (row 3~4, 병합) ──
    headers = [
        # (col, value, colspan, rowspan, bg)
        (1,  "No",               1, 2, CLR_HEADER_DARK),
        (2,  "질문 (Input)",     1, 2, CLR_HEADER_DARK),
        (3,  "CRAG 판정",        2, 1, CLR_HEADER_MID),   # 3~4
        (5,  "검색 결과 (Output)", 3, 1, CLR_HEADER_MID), # 5~7
        (8,  "응답 시간",        2, 1, CLR_HEADER_MID),   # 8~9
        (10, "LLM 생성 답변 (요약)", 2, 2, CLR_HEADER_DARK), # 10~11
    ]

    for col, val, cs, rs, bg in headers:
        style_header(ws, 3, col, val, bg=bg, colspan=cs, rowspan=rs)

    # row 4 서브헤더
    sub_headers = [
        (3, "품질 점수", CLR_HEADER_LIGHT, Font(name="맑은 고딕", bold=True, size=9, color="1F4E79")),
        (4, "판정",     CLR_HEADER_LIGHT, Font(name="맑은 고딕", bold=True, size=9, color="1F4E79")),
        (5, "Top1 정책명", CLR_HEADER_LIGHT, Font(name="맑은 고딕", bold=True, size=9, color="1F4E79")),
        (6, "Rerank Score", CLR_HEADER_LIGHT, Font(name="맑은 고딕", bold=True, size=9, color="1F4E79")),
        (7, "검색 문서 수", CLR_HEADER_LIGHT, Font(name="맑은 고딕", bold=True, size=9, color="1F4E79")),
        (8, "검색 (ms)", CLR_HEADER_LIGHT, Font(name="맑은 고딕", bold=True, size=9, color="1F4E79")),
        (9, "전체 (ms)", CLR_HEADER_LIGHT, Font(name="맑은 고딕", bold=True, size=9, color="1F4E79")),
    ]
    for col, val, bg, fnt in sub_headers:
        cell = ws.cell(row=4, column=col, value=val)
        cell.fill = fill(bg)
        cell.font = fnt
        cell.alignment = ALIGN_C
        cell.border = thin_border()

    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 20

    # ── 데이터 행 ──
    CRAG_판정_map = {
        "양호": ("품질 양호", CLR_GOOD),
        "완화": ("조건 완화 재검색", CLR_WARN),
        "폴백": ("카테고리 폴백", CLR_BAD),
    }

    # CRAG 판정 기준
    def get_crag(score):
        if score >= 0.7:   return "양호"
        elif score >= 0.4: return "완화"
        else:              return "폴백"

    # 실제 CRAG 판정 (파이프라인 출력 기준)
    actual_crag = {
        1:  ("양호",  0.924),
        2:  ("폴백",  0.306),
        3:  ("양호",  0.906),
        4:  ("양호",  0.986),
        5:  ("양호",  0.953),
        6:  ("양호",  0.846),
        7:  ("양호",  0.948),
        8:  ("완화",  0.604),
        9:  ("양호",  0.896),
        10: ("양호",  0.993),
        11: ("양호",  0.984),
        12: ("양호",  0.983),
        13: ("폴백",  0.242),
        14: ("양호",  0.848),
        15: ("양호",  0.979),
    }

    for i, r in enumerate(rows):
        row_num = i + 5
        bg = CLR_WHITE if i % 2 == 0 else CLR_GRAY
        no = r["no"]
        crag_key, crag_score = actual_crag[no]
        crag_label, crag_bg = CRAG_판정_map[crag_key]

        style_cell(ws, row_num, 1, no,               bg=bg, align=ALIGN_C, font=FONT_BODY_B)
        style_cell(ws, row_num, 2, r["question"],     bg=bg, align=ALIGN_L)
        style_cell(ws, row_num, 3, round(crag_score, 3), bg=crag_bg, align=ALIGN_C, number_format="0.000")
        style_cell(ws, row_num, 4, crag_label,        bg=crag_bg, align=ALIGN_C)
        style_cell(ws, row_num, 5, r["top1_policy"],  bg=bg, align=ALIGN_L)
        style_cell(ws, row_num, 6, r["top1_score"],   bg=bg, align=ALIGN_C, number_format="0.0000")
        style_cell(ws, row_num, 7, r["doc_count"],    bg=bg, align=ALIGN_C)
        # 첫 질문은 모델 로딩 포함 → 비고 표시
        search_ms = r["search_time_ms"]
        total_ms  = r["total_time_ms"]
        style_cell(ws, row_num, 8, search_ms, bg=bg, align=ALIGN_R, number_format="#,##0")
        style_cell(ws, row_num, 9, total_ms,  bg=bg, align=ALIGN_R, number_format="#,##0")
        # 답변 (줄여서)
        answer_short = r["answer"][:200].replace("\n", " ") if r["answer"] else "-"
        ws.merge_cells(
            start_row=row_num, start_column=10,
            end_row=row_num, end_column=11
        )
        style_cell(ws, row_num, 10, answer_short, bg=bg, align=ALIGN_L)
        ws.row_dimensions[row_num].height = 52

    # ── 비고 행 (첫 질문 검색시간 설명) ──
    note_row = len(rows) + 5
    ws.merge_cells(f"A{note_row}:K{note_row}")
    note = ws.cell(row=note_row, column=1,
                   value="※ Q01 검색시간(9,647ms)은 BGE-M3 + BM25 모델 초기 로딩 포함. 이후 평균 검색 속도: 약 160ms")
    note.font = Font(name="맑은 고딕", size=8, italic=True, color="7F7F7F")
    note.fill = fill(CLR_SECTION)
    note.alignment = ALIGN_L
    note.border = thin_border()
    ws.row_dimensions[note_row].height = 16

    # ── 컬럼 너비 ──
    col_widths = {1:5, 2:36, 3:10, 4:16, 5:28, 6:13, 7:10, 8:10, 9:10, 10:30, 11:30}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── 틀 고정 (헤더 4행 + No/질문 2열) ──
    ws.freeze_panes = "C5"


# ════════════════════════════════════════════════════════════
# 시트 2: 실험 비교 요약
# ════════════════════════════════════════════════════════════
def build_sheet2(wb):
    ws = wb.create_sheet("실험 비교 요약")
    ws.sheet_view.showGridLines = False

    # 타이틀
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "베네픽(BenePick) RAG — 실험 비교 요약"
    t.font = Font(name="맑은 고딕", bold=True, size=13, color="FFFFFF")
    t.fill = fill(CLR_HEADER_DARK)
    t.alignment = ALIGN_C
    t.border = thin_border()
    ws.row_dimensions[1].height = 28

    sections = [
        # (시작행, 섹션제목, 컬럼헤더리스트, 데이터리스트)
        (3, "① 임베딩 모델 비교 (2026-04-02 | 26개 쿼리, 1000건 샘플)",
         ["모델", "벡터 차원", "평균 품질 점수", "검색 속도(ms)", "임베딩 시간(초)", "로드 시간(초)", "비고"],
         [
             ["BAAI/bge-m3 ✅",              1024, 0.5106, 12,   506.9, 3.9,  "채택 — 현재 파이프라인"],
             ["intfloat/multilingual-e5-large", 1024, 0.4980,  8, 535.5, 86.1, "2위 — BM25 하이브리드 없음"],
             ["Alibaba-NLP/gte-Qwen2-1.5B",    1536, "N/A",  "N/A", "N/A","N/A","실행 불가 (RAM 부족, 7.1GB)"],
         ]),
        (10, "② 검색기(Sparse) 비교 (2026-04-02 | 15개 질문)",
         ["검색기", "faithfulness", "answer_relevancy", "평균", "", "", "비고"],
         [
             ["BM25 + BGE-M3 Dense ✅", 1.0000, 0.4410, 0.7205, "", "", "채택"],
             ["BGE-M3 Sparse + Dense",  0.9391, 0.4574, 0.6982, "", "", "미채택"],
         ]),
        (17, "③ Ragas 최종 평가 (2026-04-02 | 30개 질문)",
         ["지표", "점수", "", "", "", "", "비고"],
         [
             ["faithfulness",    0.9769, "", "", "", "", "환각 거의 없음"],
             ["answer_relevancy",0.5582, "", "", "", "", "개선 여지 있음"],
             ["평균",            0.7675, "", "", "", "", "전체적으로 양호"],
         ]),
        (24, "④ 파이프라인 실행 결과 요약 (2026-04-07 | 15개 질문)",
         ["항목", "값", "", "", "", "", "비고"],
         [
             ["전체 질문 수",       "15개",  "", "", "", "", ""],
             ["CRAG 품질 양호",     "12개",  "", "", "", "", "score ≥ 0.7"],
             ["CRAG 조건 완화",      "1개",  "", "", "", "", "Q08 기초수급자 (0.604)"],
             ["CRAG 카테고리 폴백",  "2개",  "", "", "", "", "Q02 훈련비(0.306), Q13 실직(0.242)"],
             ["평균 검색 속도",     "160ms", "", "", "", "", "첫 질문 모델 로딩 제외"],
             ["평균 전체 응답시간", "4.2초", "", "", "", "", "LLM 생성 포함"],
             ["Top1 평균 score",    0.9227, "", "", "", "", "Reranker 기준"],
         ]),
    ]

    for start_row, title_val, col_headers, data_rows in sections:
        # 섹션 제목
        ws.merge_cells(f"A{start_row}:G{start_row}")
        sec = ws.cell(row=start_row, column=1, value=title_val)
        sec.font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
        sec.fill = fill(CLR_HEADER_MID)
        sec.alignment = ALIGN_L
        sec.border = thin_border()
        ws.row_dimensions[start_row].height = 20

        # 컬럼 헤더
        for ci, h in enumerate(col_headers, 1):
            c = ws.cell(row=start_row+1, column=ci, value=h)
            c.font = Font(name="맑은 고딕", bold=True, size=9, color="1F4E79")
            c.fill = fill(CLR_HEADER_LIGHT)
            c.alignment = ALIGN_C
            c.border = thin_border()
        ws.row_dimensions[start_row+1].height = 18

        # 데이터
        for di, dr in enumerate(data_rows):
            row_num = start_row + 2 + di
            bg = CLR_WHITE if di % 2 == 0 else CLR_GRAY
            # 채택 행 강조
            if "채택" in str(dr[-1]):
                bg = CLR_GOOD
            for ci, val in enumerate(dr, 1):
                c = ws.cell(row=row_num, column=ci, value=val)
                c.fill = fill(bg)
                c.font = FONT_BODY
                c.alignment = ALIGN_C if ci > 1 else ALIGN_L
                c.border = thin_border()
                if isinstance(val, float) and 0 < val <= 1:
                    c.number_format = "0.0000"
            ws.row_dimensions[row_num].height = 18

    col_widths2 = {1:32, 2:16, 3:16, 4:12, 5:14, 6:14, 7:28}
    for col, w in col_widths2.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ════════════════════════════════════════════════════════════
# 메인
# ════════════════════════════════════════════════════════════
def main():
    # JSON 로드
    import glob
    json_files = sorted(glob.glob("실험결과_파이프라인_인풋아웃풋_*.json"), reverse=True)
    if not json_files:
        print("JSON 파일을 찾을 수 없습니다.")
        return
    json_path = json_files[0]
    print(f"JSON 로드: {json_path}")
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    rows = data["results"]

    wb = openpyxl.Workbook()
    build_sheet1(wb, rows)
    build_sheet2(wb)

    out_path = "실험결과_성능표_RAG_20260407.xlsx"
    wb.save(out_path)
    print(f"저장 완료: {out_path}")


if __name__ == "__main__":
    main()

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


PROJECT_ROOT = Path(__file__).resolve().parents[1]
RAG_DIR = PROJECT_ROOT / "rag"
QUESTIONS_100 = RAG_DIR / "eval_questions_100.csv"
NOTEBOOK_PATH = RAG_DIR / "compare_llm_systems_colab.ipynb"
GUIDE_PATH = RAG_DIR / "COLAB_COMPARE_GUIDE.md"
EXPERIMENT_DATASET_DIR = Path(r"C:\Users\dlfns\OneDrive\바탕 화면\실험 데이터셋\rag")
WORKBOOK_PATH = EXPERIMENT_DATASET_DIR / "09_코랩비교실험_가이드_및_로그.xlsx"


def _write_notebook() -> None:
    notebook = {
        "cells": [
            {
                "cell_type": "markdown",
                "metadata": {},
                "source": [
                    "# BenePick LLM 비교실험 (Colab)\n",
                    "\n",
                    "이 노트북은 `GPT / Claude / Gemini / BenePick` 비교 실험을 Colab에서 돌리기 위한 실행 템플릿입니다.\n",
                    "\n",
                    "비교 방식:\n",
                    "- Direct QA 비교\n",
                    "- Same Evidence QA 비교\n",
                    "- BenePick Full RAG 비교\n",
                ],
            },
            {
                "cell_type": "markdown",
                "metadata": {},
                "source": [
                    "## 1. 런타임 준비\n",
                    "- 가능하면 **GPU 런타임** 권장\n",
                    "- Google Drive에 `final_project-develope` 폴더를 올려두고 시작\n",
                ],
            },
            {
                "cell_type": "code",
                "execution_count": None,
                "metadata": {},
                "outputs": [],
                "source": [
                    "!pip install -q openai anthropic requests pandas numpy python-dotenv \\\n",
                    "    sentence-transformers transformers==4.45.2 chromadb rank-bm25 kiwipiepy \\\n",
                    "    langchain langchain-ollama langchain-groq langchain-openai langchain-community \\\n",
                    "    FlagEmbedding faiss-cpu torch\n",
                ],
            },
            {
                "cell_type": "code",
                "execution_count": None,
                "metadata": {},
                "outputs": [],
                "source": [
                    "from google.colab import drive\n",
                    "drive.mount('/content/drive')\n",
                ],
            },
            {
                "cell_type": "code",
                "execution_count": None,
                "metadata": {},
                "outputs": [],
                "source": [
                    "import os\n",
                    "PROJECT_DIR = '/content/drive/MyDrive/final_project-develope'\n",
                    "os.chdir(PROJECT_DIR)\n",
                    "print('cwd =', os.getcwd())\n",
                ],
            },
            {
                "cell_type": "markdown",
                "metadata": {},
                "source": [
                    "## 2. API 키 설정\n",
                    "Claude / Gemini는 선택입니다. 없는 경우 OpenAI + BenePick만 먼저 돌려도 됩니다.\n",
                ],
            },
            {
                "cell_type": "code",
                "execution_count": None,
                "metadata": {},
                "outputs": [],
                "source": [
                    "import os\n",
                    "\n",
                    "os.environ['OPENAI_API_KEY'] = 'YOUR_OPENAI_KEY'\n",
                    "os.environ['OPENAI_MODEL'] = 'gpt-4o-mini'\n",
                    "\n",
                    "# 선택\n",
                    "os.environ['ANTHROPIC_API_KEY'] = 'YOUR_ANTHROPIC_KEY'\n",
                    "os.environ['GEMINI_API_KEY'] = 'YOUR_GEMINI_KEY'\n",
                    "\n",
                    "# BenePick 내부 설정\n",
                    "os.environ['BENEPICK_RAG_PROMPT_VARIANT'] = 'B'\n",
                    "os.environ['BENEPICK_ENABLE_RERANKER'] = '0'\n",
                ],
            },
            {
                "cell_type": "markdown",
                "metadata": {},
                "source": [
                    "## 3. 10문항 스모크 테스트\n",
                    "비용 아끼기 위해 먼저 10문항만 돌려서 포맷/키/타임아웃 문제를 확인합니다.\n",
                ],
            },
            {
                "cell_type": "code",
                "execution_count": None,
                "metadata": {},
                "outputs": [],
                "source": [
                    "!python rag/compare_llm_systems.py \\\n",
                    "  --input rag/eval_questions_100.csv \\\n",
                    "  --limit 10 \\\n",
                    "  --targets benepick:full_rag openai:direct openai:evidence \\\n",
                    "  --output-dir rag/llm_compare_outputs_smoke\n",
                ],
            },
            {
                "cell_type": "markdown",
                "metadata": {},
                "source": [
                    "## 4. 100문항 본실험\n",
                    "OpenAI만 먼저 돌리고, 이후 Claude/Gemini를 추가하는 순서를 권장합니다.\n",
                ],
            },
            {
                "cell_type": "code",
                "execution_count": None,
                "metadata": {},
                "outputs": [],
                "source": [
                    "!python rag/compare_llm_systems.py \\\n",
                    "  --input rag/eval_questions_100.csv \\\n",
                    "  --targets benepick:full_rag openai:direct openai:evidence \\\n",
                    "  --output-dir rag/llm_compare_outputs\n",
                ],
            },
            {
                "cell_type": "code",
                "execution_count": None,
                "metadata": {},
                "outputs": [],
                "source": [
                    "# Claude / Gemini 포함 전체 비교 예시\n",
                    "!python rag/compare_llm_systems.py \\\n",
                    "  --input rag/eval_questions_100.csv \\\n",
                    "  --targets benepick:full_rag openai:direct openai:evidence claude:direct claude:evidence gemini:direct gemini:evidence \\\n",
                    "  --output-dir rag/llm_compare_outputs_all\n",
                ],
            },
            {
                "cell_type": "markdown",
                "metadata": {},
                "source": [
                    "## 5. 결과 확인\n",
                    "summary CSV에서 provider/mode별 평균을 바로 확인할 수 있습니다.\n",
                ],
            },
            {
                "cell_type": "code",
                "execution_count": None,
                "metadata": {},
                "outputs": [],
                "source": [
                    "import pandas as pd\n",
                    "from pathlib import Path\n",
                    "\n",
                    "out_dir = Path('rag/llm_compare_outputs')\n",
                    "summary_files = sorted(out_dir.glob('llm_compare_summary_*.csv'))\n",
                    "summary_files[-1] if summary_files else 'no file'\n",
                ],
            },
            {
                "cell_type": "code",
                "execution_count": None,
                "metadata": {},
                "outputs": [],
                "source": [
                    "summary = pd.read_csv(summary_files[-1])\n",
                    "summary.sort_values(['mode', 'avg_overall'], ascending=[True, False])\n",
                ],
            },
            {
                "cell_type": "markdown",
                "metadata": {},
                "source": [
                    "## 6. 권장 해석 순서\n",
                    "1. direct QA에서 범용 모델이 얼마나 강한지 본다.\n",
                    "2. evidence 기반 비교에서 생성 능력만 비교한다.\n",
                    "3. BenePick full RAG와 비교해 검색 + 생성 전체 성능을 본다.\n",
                ],
            },
        ],
        "metadata": {
            "kernelspec": {"display_name": "Python 3", "language": "python", "name": "python3"},
            "language_info": {"name": "python", "version": "3.12"},
        },
        "nbformat": 4,
        "nbformat_minor": 5,
    }
    NOTEBOOK_PATH.write_text(json.dumps(notebook, ensure_ascii=False, indent=2), encoding="utf-8")


def _write_guide() -> None:
    guide = """# BenePick LLM 비교실험 Colab 실행 가이드

## 목적
- GPT / Claude / Gemini / BenePick를 같은 질문 세트로 비교
- Direct QA / Same Evidence QA / Full RAG 3단 비교
- LLM-as-a-Judge로 groundedness, relevance, coverage, actionability, hallucination risk 평가

## 준비물
- Google Drive에 `final_project-develope` 폴더 업로드
- API 키
  - 필수: `OPENAI_API_KEY`
  - 선택: `ANTHROPIC_API_KEY`, `GEMINI_API_KEY`

## 실행 순서
1. Colab에서 `rag/compare_llm_systems_colab.ipynb` 열기
2. 패키지 설치 셀 실행
3. Drive 마운트
4. `PROJECT_DIR`를 본인 Drive 경로로 맞추기
5. API 키 입력
6. 10문항 스모크 테스트 실행
7. 이상 없으면 100문항 본실험 실행

## 추천 순서
1. `benepick:full_rag + openai:direct + openai:evidence`
2. 이후 `claude` 추가
3. 이후 `gemini` 추가

## 결과 파일
- 상세 결과: `rag/llm_compare_outputs/llm_compare_detail_*.csv`
- 요약 결과: `rag/llm_compare_outputs/llm_compare_summary_*.csv`

## 해석 포인트
- `direct`: 체감형 비교
- `evidence`: retrieval을 고정한 생성 능력 비교
- `full_rag`: BenePick 전체 파이프라인 비교
"""
    GUIDE_PATH.write_text(guide, encoding="utf-8")


def _style_header(ws, row: int, cols: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_workbook() -> None:
    EXPERIMENT_DATASET_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    questions = pd.read_csv(QUESTIONS_100)
    dist = questions["category"].value_counts().rename_axis("category").reset_index(name="count")

    storyline_rows = [
        ["1차 실험", "Query Enrichment on/off", "확장 query가 오히려 검색 정확도 저하", "enrichment 제거, query 단순화 유지"],
        ["2차 실험", "LLM-as-a-Judge v1", "groundedness는 높지만 coverage/actionability 낮음", "answer 구조화 필요 확인"],
        ["3차 실험", "구조화 답변 + guard 강화", "overall/pass_rate 개선, hallucination risk 감소", "출력 구조 통제 효과 확인"],
        ["4차 실험", "Retrieval Heuristic v1", "잡음 문서 감소, retrieval 안정성 향상", "지원정책형 질의 보정 시작"],
        ["5차 실험", "Retrieval Heuristic v2", "지역/대상자 신호 반영, 30문항 PASS 개선", "retrieval 품질 고도화"],
        ["6차 실험", "Prompt A/B 비교", "B안이 coverage/actionability 우세", "운영 후보를 B로 검토"],
        ["7차 실험", "CRAG / fallback 재조정", "broad query, 정보 부재형 질문에서 PASS 안정성 향상", "운영 안정성 확보"],
        ["8차 실험", "200문항 judge 검증", "A overall 4.61, B overall 4.71 / 둘 다 pass_rate 0.955", "B안을 운영 기본값으로 채택"],
        ["9차 예정", "GPT/Claude/Gemini 비교", "외부 범용 모델 대비 포지셔닝 검증", "100문항 비교 실험 시작"],
    ]

    compare_rows = [
        ["Direct QA", "질문만 주고 답변", "사용자 체감형 비교", "groundedness / relevance / coverage / actionability / hallucination risk"],
        ["Evidence-grounded QA", "우리 top-3 evidence 고정", "생성 모델 능력만 비교", "같은 evidence 기반 점수 비교"],
        ["Full RAG", "BenePick 전체 파이프라인", "검색+생성 전체 성능 비교", "BenePick full RAG vs 외부 모델"],
    ]

    colab_rows = [
        [1, "패키지 설치", "Colab 첫 셀에서 compare_llm_systems 실행에 필요한 패키지 설치", "!pip install -q openai anthropic requests pandas numpy python-dotenv sentence-transformers transformers==4.45.2 chromadb rank-bm25 kiwipiepy langchain langchain-ollama langchain-groq langchain-openai langchain-community FlagEmbedding faiss-cpu torch"],
        [2, "Drive 마운트", "Google Drive 연결", "from google.colab import drive; drive.mount('/content/drive')"],
        [3, "프로젝트 이동", "Drive 안 final_project-develope 경로로 이동", "PROJECT_DIR='/content/drive/MyDrive/final_project-develope'"],
        [4, "API 키 설정", "OPENAI 필수, Claude/Gemini 선택", "os.environ['OPENAI_API_KEY']='...'"],
        [5, "스모크 테스트", "10문항만 먼저 실행", "python rag/compare_llm_systems.py --input rag/eval_questions_100.csv --limit 10 --targets benepick:full_rag openai:direct openai:evidence --output-dir rag/llm_compare_outputs_smoke"],
        [6, "본실험", "100문항 비교 실행", "python rag/compare_llm_systems.py --input rag/eval_questions_100.csv --targets benepick:full_rag openai:direct openai:evidence --output-dir rag/llm_compare_outputs"],
        [7, "결과 확인", "summary CSV 확인", "rag/llm_compare_outputs/llm_compare_summary_*.csv"],
    ]

    log_columns = [
        "실험ID",
        "날짜",
        "질문세트",
        "provider",
        "mode",
        "prompt_variant",
        "num_questions",
        "avg_groundedness",
        "avg_relevance",
        "avg_coverage",
        "avg_actionability",
        "avg_hallucination_risk",
        "avg_overall",
        "pass_rate",
        "비고",
    ]

    ws = wb.create_sheet("README")
    ws["A1"] = "BenePick 비교실험 가이드 및 로그"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A3"] = "이 파일은 Colab에서 GPT / Claude / Gemini / BenePick 비교 실험을 실행하고, 결과를 누적 기록하기 위한 통합 문서입니다."
    ws["A5"] = "시트 구성"
    ws["A6"] = "- Colab_실행가이드: Colab 실행 순서"
    ws["A7"] = "- RAG_실험흐름: 1~8차 실험과 비교실험 연결"
    ws["A8"] = "- 비교실험_설계: direct / evidence / full RAG 비교 구조"
    ws["A9"] = "- 질문세트_100분포: 100문항 세트 카테고리 분포"
    ws["A10"] = "- 비교실험_결과로그: provider별 결과 기록"
    ws["A11"] = "- 실행체크리스트: 실험 전 확인 사항"
    ws.column_dimensions["A"].width = 110

    ws = wb.create_sheet("Colab_실행가이드")
    ws.append(["순서", "단계", "설명", "실행 예시"])
    _style_header(ws, 1, 4)
    for row in colab_rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 120

    ws = wb.create_sheet("RAG_실험흐름")
    ws.append(["차수", "실험명", "핵심 결과", "다음 단계 연결"])
    _style_header(ws, 1, 4)
    for row in storyline_rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 58
    ws.column_dimensions["D"].width = 38

    ws = wb.create_sheet("비교실험_설계")
    ws.append(["비교 방식", "입력", "비교 목적", "평가 항목"])
    _style_header(ws, 1, 4)
    for row in compare_rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 60

    ws = wb.create_sheet("질문세트_100분포")
    ws.append(["카테고리", "문항 수"])
    _style_header(ws, 1, 2)
    for _, row in dist.iterrows():
        ws.append([row["category"], int(row["count"])])
    ws.append([])
    ws.append(["총합", int(dist["count"].sum())])
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 10

    ws = wb.create_sheet("비교실험_결과로그")
    ws.append(log_columns)
    _style_header(ws, 1, len(log_columns))
    ws.append(["CMP-001", "", "eval_questions_100.csv", "benepick", "full_rag", "B", "", "", "", "", "", "", "", "", "예시 행"])
    for i, width in enumerate([12, 14, 24, 14, 16, 16, 14, 18, 16, 16, 18, 22, 14, 12, 28], start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    ws = wb.create_sheet("실행체크리스트")
    ws.append(["체크 항목", "설명", "확인"])
    _style_header(ws, 1, 3)
    checklist = [
        ["프로젝트 경로 확인", "Colab에서 PROJECT_DIR가 올바른지", ""],
        ["OPENAI 키 입력", "최소 OpenAI는 설정", ""],
        ["Claude/Gemini 선택 여부", "키가 있을 때만 포함", ""],
        ["10문항 스모크 테스트", "본실험 전 포맷/타임아웃 확인", ""],
        ["100문항 본실험 실행", "provider/mode별 CSV 생성 확인", ""],
        ["summary CSV 확인", "avg_overall / pass_rate 비교", ""],
        ["결과 로그 반영", "비교실험_결과로그 시트에 수동 기록", ""],
    ]
    for row in checklist:
        ws.append(row)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 10

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(WORKBOOK_PATH)


def main() -> None:
    _write_notebook()
    _write_guide()
    _write_workbook()
    print(NOTEBOOK_PATH)
    print(GUIDE_PATH)
    print(WORKBOOK_PATH)


if __name__ == "__main__":
    main()

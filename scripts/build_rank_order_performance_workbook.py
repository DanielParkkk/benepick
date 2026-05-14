from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DOWNLOADS = Path.home() / "Downloads"
SUMMARY_PATH = DOWNLOADS / "rank_order_summary_20260506_012706.csv"
DETAIL_PATH = DOWNLOADS / "rank_order_detail_20260506_012706.csv"
OUTPUT_PATH = DOWNLOADS / "BenePick_RAG_rank_order_실험성능표_20260506_012706.xlsx"

CATEGORY_KO = {
    "housing": "주거",
    "employment": "취업/일자리",
    "education": "교육/훈련",
    "health": "의료/건강",
    "basic_livelihood": "기초생활",
    "family_childcare": "가족/육아",
    "elderly_disability": "노인/장애",
    "finance_startup": "금융/창업",
}

ISSUE_KO = {
    "rank1_hit": "1위 정답",
    "in_topk_not_rank1": "Top-K 안 정답",
    "miss_topk": "Top-K 미포함",
}

DETAIL_RENAME = {
    "mode": "실험 모드",
    "row_index": "문항 번호",
    "question": "질문",
    "category": "카테고리",
    "query_type": "질문 유형",
    "expected_policy_ids": "정답 정책 ID",
    "expected_policy_names": "정답 정책명",
    "first_hit_rank": "첫 정답 순위",
    "rank_issue": "순위 판정",
    "rank_score": "순위 점수",
    "match_type": "매칭 기준",
    "matched_value": "매칭 값",
    "predicted_policy_ids_topk": "Top-K 예측 정책 ID",
    "predicted_policy_names_topk": "Top-K 예측 정책명",
    "top1_policy_name": "1위 예측",
    "top2_policy_name": "2위 예측",
    "top3_policy_name": "3위 예측",
    "latency_ms": "응답 시간(ms)",
}


def pct(value: float) -> str:
    return f"{value * 100:.1f}%"


def load_data() -> tuple[pd.DataFrame, pd.DataFrame]:
    if not SUMMARY_PATH.exists():
        raise FileNotFoundError(SUMMARY_PATH)
    if not DETAIL_PATH.exists():
        raise FileNotFoundError(DETAIL_PATH)
    summary = pd.read_csv(SUMMARY_PATH, encoding="utf-8-sig")
    detail = pd.read_csv(DETAIL_PATH, encoding="utf-8-sig")
    detail["category_ko"] = detail["category"].map(CATEGORY_KO).fillna(detail["category"])
    detail["rank_issue_ko"] = detail["rank_issue"].map(ISSUE_KO).fillna(detail["rank_issue"])
    return summary, detail


def build_derived_tables(summary: pd.DataFrame, detail: pd.DataFrame) -> dict[str, pd.DataFrame]:
    category = (
        detail.groupby(["mode", "category", "category_ko"], dropna=False)
        .agg(
            num_questions=("question", "count"),
            hit_at_1=("first_hit_rank", lambda s: (s == 1).mean()),
            hit_at_3=("first_hit_rank", lambda s: ((s.notna()) & (s <= 3)).mean()),
            hit_at_5=("first_hit_rank", lambda s: ((s.notna()) & (s <= 5)).mean()),
            avg_rank_score=("rank_score", "mean"),
            avg_latency_ms=("latency_ms", "mean"),
        )
        .reset_index()
        .sort_values(["hit_at_1", "avg_rank_score"], ascending=[False, False])
    )

    issue = (
        detail.groupby(["mode", "rank_issue", "rank_issue_ko"], dropna=False)
        .size()
        .reset_index(name="count")
        .sort_values("count", ascending=False)
    )
    issue["rate"] = issue["count"] / max(1, len(detail))

    misses = detail[detail["rank_issue"].eq("miss_topk")].copy()
    weak = detail[detail["rank_issue"].ne("rank1_hit")].copy()
    weak = weak.sort_values(["rank_issue", "first_hit_rank", "latency_ms"], ascending=[True, True, False])

    return {
        "category": category,
        "issue": issue,
        "misses": misses,
        "weak": weak,
    }


def setup_sheet(ws, title: str, subtitle: str | None = None) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=18, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:H1")
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(size=10, color="666666")
        ws.merge_cells("A2:H2")


def style_table(ws, header_row: int, max_row: int, max_col: int) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    border = Border(bottom=Side(style="thin", color="B7C9D6"))
    for cell in ws[header_row]:
        if cell.column <= max_col:
            cell.font = Font(bold=True, color="1F4E79")
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=header_row + 1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="E6EEF4"))


def write_df(ws, df: pd.DataFrame, start_row: int, start_col: int = 1, index: bool = False) -> tuple[int, int]:
    out = df.reset_index() if index else df
    for c_idx, col in enumerate(out.columns, start_col):
        ws.cell(start_row, c_idx, col)
    for r_offset, row in enumerate(out.itertuples(index=False), 1):
        for c_offset, value in enumerate(row, 0):
            if pd.isna(value):
                value = None
            ws.cell(start_row + r_offset, start_col + c_offset, value)
    max_row = start_row + len(out)
    max_col = start_col + len(out.columns) - 1
    style_table(ws, start_row, max_row, max_col)
    return max_row, max_col


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_dashboard(wb: Workbook, summary: pd.DataFrame, derived: dict[str, pd.DataFrame]) -> None:
    ws = wb.active
    ws.title = "실험 성능표"
    row = summary.iloc[0]
    setup_sheet(
        ws,
        "BenePick RAG Rank Order 실험 성능표",
        f"실험 시각: {row['timestamp']} | 평가셋: {row['labels_path']} | 생성: {datetime.now():%Y-%m-%d %H:%M}",
    )

    kpis = [
        ("평가 문항", int(row["num_questions"]), "라벨 있는 문항 기준"),
        ("Top-1 Hit", row["hit_at_1"], "1위 결과가 정답"),
        ("Top-3 Hit", row["hit_at_3"], "3위 안 정답 포함"),
        ("Top-5 Hit", row["hit_at_5"], "5위 안 정답 포함"),
        ("MRR", row["mrr"], "정답 순위 역수 평균"),
        ("평균 순위점수", row["avg_rank_score"], "rank_score 평균"),
        ("평균 지연", row["avg_latency_ms"], "ms"),
        ("커버리지", row["coverage_rate"], "라벨 커버리지"),
    ]
    ws["A4"] = "핵심 KPI"
    ws["A4"].font = Font(bold=True, size=13, color="1F4E79")
    for idx, (label, value, note) in enumerate(kpis):
        r = 5 + idx // 4 * 4
        c = 1 + idx % 4 * 2
        ws.cell(r, c, label)
        ws.cell(r + 1, c, value)
        ws.cell(r + 2, c, note)
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
        ws.merge_cells(start_row=r + 1, start_column=c, end_row=r + 1, end_column=c + 1)
        ws.merge_cells(start_row=r + 2, start_column=c, end_row=r + 2, end_column=c + 1)
        for rr in range(r, r + 3):
            for cc in range(c, c + 2):
                ws.cell(rr, cc).fill = PatternFill("solid", fgColor="F4F8FB")
        ws.cell(r, c).font = Font(bold=True, color="5B6770")
        ws.cell(r + 1, c).font = Font(bold=True, size=16, color="0F5B8C")
        ws.cell(r + 2, c).font = Font(size=9, color="666666")
        if isinstance(value, float) and label not in {"평균 지연"}:
            ws.cell(r + 1, c).number_format = "0.0%"
        elif label == "평균 지연":
            ws.cell(r + 1, c).number_format = "0.0"

    ws["A14"] = "해석 요약"
    ws["A14"].font = Font(bold=True, size=13, color="1F4E79")
    issue = derived["issue"]
    miss_count = int(issue.loc[issue["rank_issue"].eq("miss_topk"), "count"].sum())
    topk_not_rank1 = int(issue.loc[issue["rank_issue"].eq("in_topk_not_rank1"), "count"].sum())
    bullets = [
        f"Top-5 Hit {pct(row['hit_at_5'])}, Top-3 Hit {pct(row['hit_at_3'])}로 검색 후보군 안에는 정답이 대부분 포함됩니다.",
        f"Top-1 Hit은 {pct(row['hit_at_1'])}입니다. 정답은 찾지만 1위 정렬에서 밀린 케이스가 {topk_not_rank1}건입니다.",
        f"Top-K 미포함은 {miss_count}건입니다. 이 문항들은 검색 recall 자체 보강이 필요합니다.",
        f"평균 응답 시간은 {row['avg_latency_ms']:.1f}ms로, rank-order 평가 기준에서는 실사용 가능 범위입니다.",
    ]
    for i, text in enumerate(bullets, 15):
        ws.cell(i, 1, f"- {text}")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
        ws.cell(i, 1).alignment = Alignment(wrap_text=True, vertical="top")

    metric_table = pd.DataFrame(
        [
            ["Top-1 Hit", row["hit_at_1"]],
            ["Top-3 Hit", row["hit_at_3"]],
            ["Top-5 Hit", row["hit_at_5"]],
            ["MRR", row["mrr"]],
            ["평균 순위점수", row["avg_rank_score"]],
        ],
        columns=["지표", "값"],
    )
    write_df(ws, metric_table, 21, 1)
    for r in range(22, 27):
        ws.cell(r, 2).number_format = "0.0%"

    chart = BarChart()
    chart.title = "Rank Order 주요 지표"
    chart.y_axis.title = "Score"
    chart.x_axis.title = "Metric"
    chart.height = 7
    chart.width = 15
    chart.add_data(Reference(ws, min_col=2, min_row=21, max_row=26), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=22, max_row=26))
    ws.add_chart(chart, "D21")

    issue_display = derived["issue"][["rank_issue_ko", "count", "rate"]].rename(
        columns={"rank_issue_ko": "순위 판정", "count": "건수", "rate": "비율"}
    )
    write_df(ws, issue_display, 31, 1)
    for r in range(32, 32 + len(issue_display)):
        ws.cell(r, 3).number_format = "0.0%"

    issue_chart = BarChart()
    issue_chart.title = "순위 판정 분포"
    issue_chart.height = 7
    issue_chart.width = 15
    issue_chart.add_data(Reference(ws, min_col=2, min_row=31, max_row=31 + len(issue_display)), titles_from_data=True)
    issue_chart.set_categories(Reference(ws, min_col=1, min_row=32, max_row=31 + len(issue_display)))
    ws.add_chart(issue_chart, "D31")

    set_widths(ws, {"A": 18, "B": 14, "C": 14, "D": 16, "E": 14, "F": 14, "G": 14, "H": 14})
    ws.freeze_panes = "A4"


def add_category_sheet(wb: Workbook, derived: dict[str, pd.DataFrame]) -> None:
    ws = wb.create_sheet("카테고리별 성능")
    setup_sheet(ws, "카테고리별 성능", "카테고리별 Hit@K, 평균 순위점수, 지연시간")
    category = derived["category"].copy()
    category = category.rename(
        columns={
            "mode": "실험 모드",
            "category": "카테고리 코드",
            "category_ko": "카테고리",
            "num_questions": "문항 수",
            "hit_at_1": "Hit@1",
            "hit_at_3": "Hit@3",
            "hit_at_5": "Hit@5",
            "avg_rank_score": "평균 순위점수",
            "avg_latency_ms": "평균 지연(ms)",
        }
    )
    write_df(ws, category, 4)
    for row in range(5, 5 + len(category)):
        for col in [5, 6, 7, 8]:
            ws.cell(row, col).number_format = "0.0%"
        ws.cell(row, 9).number_format = "0.0"

    chart = BarChart()
    chart.title = "카테고리별 Hit@1"
    chart.height = 8
    chart.width = 18
    chart.add_data(Reference(ws, min_col=5, min_row=4, max_row=4 + len(category)), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=3, min_row=5, max_row=4 + len(category)))
    ws.add_chart(chart, "K4")

    line = LineChart()
    line.title = "카테고리별 Hit@3 / Hit@5"
    line.height = 8
    line.width = 18
    line.add_data(Reference(ws, min_col=6, max_col=7, min_row=4, max_row=4 + len(category)), titles_from_data=True)
    line.set_categories(Reference(ws, min_col=3, min_row=5, max_row=4 + len(category)))
    ws.add_chart(line, "K20")
    set_widths(ws, {"A": 14, "B": 18, "C": 16, "D": 10, "E": 10, "F": 10, "G": 10, "H": 14, "I": 14, "K": 20})
    ws.freeze_panes = "A5"


def add_detail_sheet(wb: Workbook, detail: pd.DataFrame) -> None:
    ws = wb.create_sheet("상세 결과")
    setup_sheet(ws, "문항별 상세 결과", "각 질문별 정답 정책, Top-K 예측, 순위 판정")
    display = detail.drop(columns=["category_ko", "rank_issue_ko"]).rename(columns=DETAIL_RENAME)
    display["카테고리"] = detail["category_ko"]
    display["순위 판정"] = detail["rank_issue_ko"]
    ordered = [
        "문항 번호", "카테고리", "질문", "정답 정책명", "첫 정답 순위", "순위 판정",
        "순위 점수", "1위 예측", "2위 예측", "3위 예측", "응답 시간(ms)",
        "정답 정책 ID", "Top-K 예측 정책명", "Top-K 예측 정책 ID", "매칭 기준", "매칭 값",
    ]
    display = display[[col for col in ordered if col in display.columns]]
    write_df(ws, display, 4)
    for row in range(5, 5 + len(display)):
        ws.cell(row, 7).number_format = "0.0"
        ws.cell(row, 11).number_format = "0.0"
        issue = ws.cell(row, 6).value
        fill = None
        if issue == "1위 정답":
            fill = PatternFill("solid", fgColor="E2F0D9")
        elif issue == "Top-K 안 정답":
            fill = PatternFill("solid", fgColor="FFF2CC")
        elif issue == "Top-K 미포함":
            fill = PatternFill("solid", fgColor="FCE4D6")
        if fill:
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = fill
    set_widths(
        ws,
        {
            "A": 10, "B": 14, "C": 42, "D": 38, "E": 12, "F": 14,
            "G": 10, "H": 36, "I": 36, "J": 36, "K": 14, "L": 28,
            "M": 60, "N": 50, "O": 12, "P": 28,
        },
    )
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(ws.max_column)}{ws.max_row}"


def add_error_sheet(wb: Workbook, derived: dict[str, pd.DataFrame]) -> None:
    ws = wb.create_sheet("개선 필요 문항")
    setup_sheet(ws, "개선 필요 문항", "Top-K 미포함 및 1위 정렬 실패 케이스")
    weak = derived["weak"].copy()
    weak["카테고리"] = weak["category_ko"]
    weak["순위 판정"] = weak["rank_issue_ko"]
    weak = weak.rename(columns=DETAIL_RENAME)
    cols = ["문항 번호", "카테고리", "질문", "순위 판정", "첫 정답 순위", "정답 정책명", "1위 예측", "2위 예측", "3위 예측", "순위 점수", "응답 시간(ms)"]
    weak = weak[[c for c in cols if c in weak.columns]]
    write_df(ws, weak, 4)
    for row in range(5, 5 + len(weak)):
        ws.cell(row, 10).number_format = "0.0"
        ws.cell(row, 11).number_format = "0.0"
    set_widths(ws, {"A": 10, "B": 14, "C": 44, "D": 15, "E": 12, "F": 42, "G": 38, "H": 38, "I": 38, "J": 10, "K": 14})
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(ws.max_column)}{ws.max_row}"


def add_source_sheets(wb: Workbook, summary: pd.DataFrame, detail: pd.DataFrame, derived: dict[str, pd.DataFrame]) -> None:
    ws = wb.create_sheet("원본 Summary")
    setup_sheet(ws, "원본 Summary", "실험 원본 요약 데이터")
    write_df(ws, summary, 4)
    set_widths(ws, {get_column_letter(i): 16 for i in range(1, ws.max_column + 1)})

    ws2 = wb.create_sheet("원본 Detail")
    setup_sheet(ws2, "원본 Detail", "실험 원본 상세 데이터")
    write_df(ws2, detail.drop(columns=["category_ko", "rank_issue_ko"]), 4)
    set_widths(ws2, {get_column_letter(i): 18 for i in range(1, ws2.max_column + 1)})
    ws2.column_dimensions["C"].width = 42
    ws2.column_dimensions["G"].width = 45
    ws2.freeze_panes = "A5"
    ws2.auto_filter.ref = f"A4:{get_column_letter(ws2.max_column)}{ws2.max_row}"


def finalize(wb: Workbook) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal or "left",
                    vertical=cell.alignment.vertical or "center",
                    wrap_text=cell.alignment.wrap_text,
                )
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0


def main() -> None:
    summary, detail = load_data()
    derived = build_derived_tables(summary, detail)
    wb = Workbook()
    add_dashboard(wb, summary, derived)
    add_category_sheet(wb, derived)
    add_error_sheet(wb, derived)
    add_detail_sheet(wb, detail)
    add_source_sheets(wb, summary, detail, derived)
    finalize(wb)
    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()

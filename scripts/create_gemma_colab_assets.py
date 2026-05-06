from __future__ import annotations

import json
import shutil
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path("D:/Projects/final_project-develope")
DOWNLOADS = Path("C:/Users/dlfns/Downloads")
BUNDLE_NAME = "13_Gemma_Ollama_RAG_experiment_bundle.zip"
NOTEBOOK_NAME = "gemma_ollama_rag_experiment_colab.ipynb"
LOG_NAME = "13_Gemma_Ollama_RAG_experiment_log.xlsx"


def cell(source: str, cell_type: str = "code") -> dict:
    if cell_type == "markdown":
        return {"cell_type": "markdown", "metadata": {}, "source": source.splitlines(keepends=True)}
    return {
        "cell_type": "code",
        "execution_count": None,
        "metadata": {},
        "outputs": [],
        "source": source.splitlines(keepends=True),
    }


def create_notebook(path: Path) -> None:
    cells = [
        cell(
            """# BenePick Gemma/Ollama RAG Experiment

목표:
- Qwen3.5:4b와 Gemma 4를 Ollama 로컬 실행 방식으로 비교합니다.
- 기존 BGE-M3 임베딩과 EmbeddingGemma 임베딩의 검색 순위 성능을 비교합니다.
- 결과 CSV/XLSX는 Drive의 `BenePick_gemma_experiment_outputs` 폴더로 복사합니다.

주의:
- 코랩 런타임 경로는 한글 깨짐을 피하려고 `/content/benepick_gemma_experiment`로 고정합니다.
- Gemma 4 기본 태그는 `gemma4:e2b`입니다. 더 좋은 모델을 쓰려면 `gemma4:e4b`, `gemma4:26b`로 바꿔 재실험하세요.
- EmbeddingGemma는 Hugging Face에서 라이선스 동의 후 `HF_TOKEN`이 필요합니다.
""",
            "markdown",
        ),
        cell(
            """# 1. Mount Drive and extract bundle
from google.colab import drive
drive.mount('/content/drive')

import os
import shutil
import zipfile
from pathlib import Path

DRIVE_ROOT = Path('/content/drive/MyDrive')
BUNDLE_ZIP = DRIVE_ROOT / '13_Gemma_Ollama_RAG_experiment_bundle.zip'
RUNTIME_ROOT = Path('/content/benepick_gemma_experiment')

print('bundle exists:', BUNDLE_ZIP.exists(), BUNDLE_ZIP)
if not BUNDLE_ZIP.exists():
    raise FileNotFoundError('Upload 13_Gemma_Ollama_RAG_experiment_bundle.zip to MyDrive first.')

if RUNTIME_ROOT.exists():
    shutil.rmtree(RUNTIME_ROOT)
RUNTIME_ROOT.mkdir(parents=True, exist_ok=True)

with zipfile.ZipFile(BUNDLE_ZIP, 'r') as zf:
    zf.extractall(RUNTIME_ROOT)

candidates = sorted(RUNTIME_ROOT.rglob('rag/compare_ollama_generation.py'))
if not candidates:
    raise FileNotFoundError('rag/compare_ollama_generation.py not found after extraction.')

PROJECT_DIR = candidates[0].parents[1]
os.chdir(PROJECT_DIR)
print('PROJECT_DIR =', PROJECT_DIR)
print('cwd =', os.getcwd())
print('labels =', Path('rag/eval_labels_template_100_filled.csv').exists())
""",
        ),
        cell(
            """# 2. Install Python requirements
import subprocess, sys
from pathlib import Path

REQ_FILE = Path('requirements_colab_gemma.txt')
print('req exists:', REQ_FILE.exists(), REQ_FILE)
subprocess.run([sys.executable, '-m', 'pip', 'install', '-q', '-r', str(REQ_FILE)], check=True)
print('install done')
""",
        ),
        cell(
            """# 3. Optional: set secrets for judge and EmbeddingGemma
# OpenAI key is only needed if you want LLM-as-a-Judge scores for Qwen vs Gemma generation.
# HF_TOKEN is needed only for google/embeddinggemma-300m after accepting the Hugging Face license.
import os

try:
    from google.colab import userdata
    os.environ['OPENAI_API_KEY'] = userdata.get('OPENAI_API_KEY') or userdata.get('GPT_API_KEY') or os.environ.get('OPENAI_API_KEY', '')
    os.environ['HF_TOKEN'] = userdata.get('HF_TOKEN') or userdata.get('HUGGINGFACE_HUB_TOKEN') or os.environ.get('HF_TOKEN', '')
except Exception as exc:
    print('Colab userdata not available:', exc)

print('OPENAI_API_KEY set:', bool(os.environ.get('OPENAI_API_KEY')))
print('HF_TOKEN set:', bool(os.environ.get('HF_TOKEN')))
""",
        ),
        cell(
            """# 4. Install and start Ollama
import shutil
import subprocess
import time
import requests
import os

def install_ollama_colab():
    if shutil.which('ollama'):
        print('ollama already installed:', shutil.which('ollama'))
        return

    # First try the official one-line installer. In some Colab runtimes this
    # can fail while trying to configure service users, so we keep a fallback.
    print('trying official install.sh...')
    first = subprocess.run(
        'curl -fsSL https://ollama.com/install.sh | sh',
        shell=True,
        text=True,
        capture_output=True,
    )
    if first.returncode == 0 and shutil.which('ollama'):
        print('official install complete')
        return

    print('official installer failed; using manual tar.zst install')
    print('stdout tail:', first.stdout[-1000:])
    print('stderr tail:', first.stderr[-1000:])

    subprocess.run(['apt-get', '-qq', 'update'], check=True)
    subprocess.run(['apt-get', '-qq', 'install', '-y', 'zstd'], check=True)
    subprocess.run(
        'curl -fsSL https://ollama.com/download/ollama-linux-amd64.tar.zst | tar -x --zstd -C /usr',
        shell=True,
        check=True,
    )

    if not shutil.which('ollama'):
        raise RuntimeError('Ollama install finished but ollama command was not found in PATH')

install_ollama_colab()
subprocess.run(['ollama', '-v'], check=False)

OLLAMA_HOST = '127.0.0.1:11434'
os.environ['OLLAMA_HOST'] = OLLAMA_HOST
ollama_proc = subprocess.Popen(['ollama', 'serve'], stdout=subprocess.DEVNULL, stderr=subprocess.STDOUT)
time.sleep(8)

for i in range(20):
    try:
        r = requests.get('http://127.0.0.1:11434/api/tags', timeout=5)
        print('ollama ready:', r.status_code)
        break
    except Exception as exc:
        if i == 19:
            raise
        time.sleep(2)
""",
        ),
        cell(
            """# 5. Pull local generation models
import subprocess

QWEN_MODEL = 'qwen3.5:4b'
GEMMA_MODEL = 'gemma4:e2b'  # Try gemma4:e4b or gemma4:26b if runtime allows.

for model in [QWEN_MODEL, GEMMA_MODEL]:
    print('pulling:', model)
    subprocess.run(['ollama', 'pull', model], check=True)

subprocess.run(['ollama', 'list'], check=True)
""",
        ),
        cell(
            """# 6. Smoke test: Qwen vs Gemma generation, no judge
import subprocess, sys

cmd = [
    sys.executable, '-u', 'rag/compare_ollama_generation.py',
    '--input', 'rag/eval_questions_100.csv',
    '--output-dir', 'rag/ollama_generation_compare_smoke',
    '--models', f'qwen={QWEN_MODEL}', f'gemma={GEMMA_MODEL}',
    '--modes', 'direct', 'evidence',
    '--limit', '3',
    '--judge', 'none',
    '--num-predict', '220',
    '--timeout', '120',
]
print('Running:', ' '.join(cmd))
subprocess.run(cmd, check=True)
""",
        ),
        cell(
            """# 7. Full generation comparison: Qwen vs Gemma
# If OPENAI_API_KEY is set, this also creates LLM-as-a-Judge scores.
import os, subprocess, sys

judge_mode = 'openai' if os.environ.get('OPENAI_API_KEY') else 'none'
cmd = [
    sys.executable, '-u', 'rag/compare_ollama_generation.py',
    '--input', 'rag/eval_questions_100.csv',
    '--output-dir', 'rag/ollama_generation_compare_full',
    '--models', f'qwen={QWEN_MODEL}', f'gemma={GEMMA_MODEL}',
    '--modes', 'direct', 'evidence',
    '--limit', '100',
    '--judge', judge_mode,
    '--num-predict', '220',
    '--timeout', '120',
]
print('judge_mode =', judge_mode)
print('Running:', ' '.join(cmd))
subprocess.run(cmd, check=True)
""",
        ),
        cell(
            """# 8. BGE-M3 rank-order baseline
import os, subprocess, sys

env = os.environ.copy()
env['BENEPICK_SKIP_CHROMA'] = '1'
env['BENEPICK_EMBED_MODEL'] = 'BAAI/bge-m3'
env.pop('BENEPICK_WELFARE_EMBEDDINGS_PATH', None)
env.pop('BENEPICK_GOV24_EMBEDDINGS_PATH', None)

cmd = [
    sys.executable, '-u', 'rag/evaluate_rank_order_colab_v2.py',
    '--labels', 'rag/eval_labels_template_100_filled.csv',
    '--output-dir', 'rag/rank_order_bge_m3',
    '--expected-min-labels', '90',
]
print('Running:', ' '.join(cmd))
subprocess.run(cmd, check=True, env=env)
""",
        ),
        cell(
            """# 9. Build EmbeddingGemma embeddings
# Run this only after accepting the model license on Hugging Face and setting HF_TOKEN in Colab secrets.
import os, subprocess, sys

if not os.environ.get('HF_TOKEN'):
    raise RuntimeError('HF_TOKEN is not set. Accept google/embeddinggemma-300m license on Hugging Face and add HF_TOKEN to Colab secrets.')

cmd = [
    sys.executable, '-u', 'rag/build_embedding_variants.py',
    '--model', 'google/embeddinggemma-300m',
    '--device', 'cuda',
    '--batch-size', '32',
]
print('Running:', ' '.join(cmd))
subprocess.run(cmd, check=True)
""",
        ),
        cell(
            """# 10. Compare BGE-M3 vs EmbeddingGemma rank order
import os, subprocess, sys

env = os.environ.copy()
env['BENEPICK_SKIP_CHROMA'] = '1'
env['HF_TOKEN'] = os.environ.get('HF_TOKEN', '')

cmd = [
    sys.executable, '-u', 'rag/compare_embedding_rank_order.py',
    '--labels', 'rag/eval_labels_template_100_filled.csv',
    '--expected-min-labels', '90',
]
print('Running:', ' '.join(cmd))
subprocess.run(cmd, check=True, env=env)
""",
        ),
        cell(
            """# 11. Collect latest results and copy to Drive
import shutil
from pathlib import Path

EXPORT_DIR = DRIVE_ROOT / 'BenePick_gemma_experiment_outputs'
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

patterns = [
    'rag/ollama_generation_compare_smoke/*',
    'rag/ollama_generation_compare_full/*',
    'rag/rank_order_bge_m3/*',
    'rag/rank_order_embedding_compare/*',
]

copied = []
for pattern in patterns:
    for file in Path('.').glob(pattern):
        if file.is_file() and file.suffix.lower() in {'.csv', '.xlsx'}:
            target = EXPORT_DIR / file.name
            shutil.copy2(file, target)
            copied.append(target)

print('copied files:')
for item in copied:
    print(item)
""",
        ),
    ]
    notebook = {
        "cells": cells,
        "metadata": {
            "kernelspec": {"display_name": "Python 3", "language": "python", "name": "python3"},
            "language_info": {"name": "python", "version": "3.x"},
        },
        "nbformat": 4,
        "nbformat_minor": 5,
    }
    path.write_text(json.dumps(notebook, ensure_ascii=False, indent=2), encoding="utf-8")


def add_table(ws, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell_obj in ws[1]:
        cell_obj.font = Font(bold=True, color="FFFFFF")
        cell_obj.fill = header_fill
        cell_obj.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell_obj in row:
            cell_obj.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [18, 24, 22, 22, 14, 16, 18, 14, 14, 14, 14, 14, 14, 14, 20, 42]
    for i, width in enumerate(widths[: ws.max_column], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_log(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "실험_개요"
    overview = [
        ["항목", "내용"],
        ["목표", "Gemma 4 기반 BenePick RAG가 기존 Qwen/Ollama 및 BGE-M3 임베딩 대비 어느 정도 성능인지 확인"],
        ["대회 전략", "Ollama 특별상 1순위: 민감한 복지 조건을 로컬 우선으로 처리하는 Gemma 4 RAG"],
        ["생성 모델 비교", "Qwen3.5:4b vs Gemma4:e2b direct/evidence 답변 비교"],
        ["검색 임베딩 비교", "BGE-M3 baseline vs EmbeddingGemma rank-order Hit@K/MRR 비교"],
        ["기존 기준값", "BGE-M3 97문항 라벨 기준 Hit@1 0.6907, Hit@3 0.8866, Hit@5 0.9485, MRR 0.7933"],
        ["판단 기준", "답변 품질은 LLM-as-a-Judge, 정렬 순위는 라벨 기반 Hit@K/MRR로 분리 평가"],
        ["주의", "EmbeddingGemma는 Hugging Face 라이선스 동의와 HF_TOKEN이 필요"],
    ]
    for row in overview:
        ws.append(row)
    for cell_obj in ws[1]:
        cell_obj.font = Font(bold=True, color="FFFFFF")
        cell_obj.fill = PatternFill("solid", fgColor="1F4E78")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 100
    for row in ws.iter_rows():
        for cell_obj in row:
            cell_obj.alignment = Alignment(vertical="top", wrap_text=True)

    ws = wb.create_sheet("생성모델_Qwen_vs_Gemma")
    add_table(
        ws,
        [
            "실험번호",
            "비교대상",
            "모델_A",
            "모델_B",
            "질의수",
            "모드",
            "Judge",
            "A_overall",
            "B_overall",
            "A_pass_rate",
            "B_pass_rate",
            "A_환각위험",
            "B_환각위험",
            "채택",
            "결론",
            "비고",
        ],
        [
            [
                "13-1",
                "Ollama 생성 모델",
                "qwen3.5:4b",
                "gemma4:e2b",
                100,
                "direct",
                "OpenAI judge 또는 수동검토",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "일반 LLM 직답에서 정책 조건을 얼마나 정확히 말하는지 비교",
                "코랩 7번 셀 결과 입력",
            ],
            [
                "13-2",
                "Ollama 생성 모델",
                "qwen3.5:4b",
                "gemma4:e2b",
                100,
                "evidence",
                "OpenAI judge 또는 수동검토",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "같은 top-3 근거를 줬을 때 Gemma가 답변을 더 잘 구조화하는지 비교",
                "코랩 7번 셀 결과 입력",
            ],
        ],
    )

    ws = wb.create_sheet("임베딩_BGE_vs_EmbeddingGemma")
    add_table(
        ws,
        [
            "실험번호",
            "비교대상",
            "임베딩_A",
            "임베딩_B",
            "라벨문항수",
            "라벨커버리지",
            "A_Hit@1",
            "B_Hit@1",
            "A_Hit@3",
            "B_Hit@3",
            "A_Hit@5",
            "B_Hit@5",
            "A_MRR",
            "B_MRR",
            "채택",
            "비고",
        ],
        [
            [
                "13-3",
                "검색 정렬 임베딩",
                "BAAI/bge-m3",
                "google/embeddinggemma-300m",
                97,
                0.97,
                0.6907,
                "",
                0.8866,
                "",
                0.9485,
                "",
                0.7933,
                "",
                "",
                "BGE-M3 기존 기준값은 97문항 라벨 기반 결과. EmbeddingGemma 결과는 코랩 10번 셀 후 입력",
            ]
        ],
    )

    ws = wb.create_sheet("Ollama_특별상_체크리스트")
    add_table(
        ws,
        ["체크항목", "상태", "설명", "발표에서 말할 문장"],
        [
            ["Gemma 4 핵심 사용", "진행", "답변 생성/요약/번역 기본 모델을 Gemma 4로 전환", "BenePick의 핵심 생성 모델을 Gemma 4로 교체했습니다."],
            ["Ollama 로컬 실행", "진행", "코랩과 로컬에서 Ollama API로 Gemma 4 실행", "민감한 복지 조건을 외부 API에 보내지 않는 로컬 우선 구조를 실험했습니다."],
            ["BGE-M3 대비 EmbeddingGemma", "실험 예정", "Hit@K/MRR로 검색 순위 비교", "생성 모델뿐 아니라 검색 임베딩도 Gemma 계열로 대체 가능한지 검증했습니다."],
            ["재현 가능성", "준비", "노트북, 번들, 실행 명령, 결과표 제공", "한 번의 코랩 실행으로 실험을 재현할 수 있게 구성했습니다."],
            ["CC-BY 4.0 준비", "확인 필요", "수상 시 코드/설명을 공개해야 함", "대회 제출 전 공개 가능한 코드와 비공개 키/데이터를 분리했습니다."],
        ],
    )

    ws = wb.create_sheet("실행_명령어")
    add_table(
        ws,
        ["단계", "명령어", "결과파일", "메모"],
        [
            ["생성 smoke", "python rag/compare_ollama_generation.py --models qwen=qwen3.5:4b gemma=gemma4:e2b --limit 3 --judge none", "ollama_generation_compare_smoke", "모델 호출/경로 확인"],
            ["생성 full", "python rag/compare_ollama_generation.py --models qwen=qwen3.5:4b gemma=gemma4:e2b --limit 100 --judge openai", "ollama_generation_compare_full", "OPENAI_API_KEY가 있으면 judge 점수 생성"],
            ["BGE baseline", "python rag/evaluate_rank_order_colab_v2.py --labels rag/eval_labels_template_100_filled.csv --output-dir rag/rank_order_bge_m3 --expected-min-labels 90", "rank_order_bge_m3", "기존 검색 기준값 재현"],
            ["EmbeddingGemma build", "python rag/build_embedding_variants.py --model google/embeddinggemma-300m --device cuda --batch-size 32", "processed/*embeddinggemma*.npy", "HF_TOKEN 필요"],
            ["Embedding compare", "python rag/compare_embedding_rank_order.py --labels rag/eval_labels_template_100_filled.csv --expected-min-labels 90", "rank_order_embedding_compare", "BGE-M3 vs EmbeddingGemma 비교"],
        ],
    )

    wb.save(path)


def add_file_to_zip(zf: zipfile.ZipFile, file: Path, root: Path) -> None:
    arcname = file.relative_to(root).as_posix()
    zf.write(file, arcname)


def create_bundle(path: Path) -> None:
    include_files = [
        "requirements_colab_gemma.txt",
        "rag/__init__.py",
        "rag/searcher.py",
        "rag/evaluate_rank_order_colab_v2.py",
        "rag/build_embedding_variants.py",
        "rag/compare_embedding_rank_order.py",
        "rag/compare_ollama_generation.py",
        "rag/eval_questions_100.csv",
        "rag/eval_labels_template_100_filled.csv",
        "rag/GEMMA_EXPERIMENT_GUIDE.md",
        "processed/chunks.csv",
        "processed/embeddings.npy",
        "processed/bm25_cache.pkl",
        "processed/bm25_cache.hash",
        "processed/gov24/chunks.csv",
        "processed/gov24/embeddings.npy",
    ]
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        for rel in include_files:
            file = ROOT / rel
            if not file.exists():
                raise FileNotFoundError(file)
            add_file_to_zip(zf, file, ROOT)


def verify_no_question_marks(path: Path) -> None:
    if path.suffix == ".ipynb":
        text = path.read_text(encoding="utf-8")
        broken = text.count("????")
        if broken:
            raise RuntimeError(f"Notebook may contain broken text: {path} count={broken}")
    if path.suffix == ".xlsx":
        # The binary xlsx may naturally contain question mark bytes, so verify by re-opening cells.
        from openpyxl import load_workbook

        wb = load_workbook(path)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and "????" in c.value:
                        raise RuntimeError(f"Broken text found in {path}:{ws.title}!{c.coordinate}")


def main() -> None:
    DOWNLOADS.mkdir(parents=True, exist_ok=True)
    notebook_path = DOWNLOADS / NOTEBOOK_NAME
    log_path = DOWNLOADS / LOG_NAME
    bundle_path = DOWNLOADS / BUNDLE_NAME

    create_notebook(notebook_path)
    create_log(log_path)
    create_bundle(bundle_path)

    verify_no_question_marks(notebook_path)
    verify_no_question_marks(log_path)

    # Also keep a copy of the notebook in the repo for version control.
    shutil.copy2(notebook_path, ROOT / "rag" / NOTEBOOK_NAME)

    print("created notebook:", notebook_path)
    print("created bundle:", bundle_path)
    print("created log:", log_path)


if __name__ == "__main__":
    main()
